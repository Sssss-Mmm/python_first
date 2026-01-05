"""
[Step 03] Excel Automation (openpyxl)
--------------------------------------
RPA 업무의 80%는 엑셀 작업입니다.
Python의 'openpyxl' 라이브러리를 사용하면 엑셀을 직접 켜지 않고도
데이터를 읽고, 쓰고, 서식을 변경할 수 있습니다.

[필수 설치]
pip install openpyxl
"""

import os
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("[Error] 'openpyxl' 라이브러리가 설치되지 않았습니다.")
    print("터미널에서 'pip install openpyxl'을 실행해주세요.")
    exit()

def create_excel_report(filename):
    """
    새로운 엑셀 파일을 생성하고 데이터를 입력합니다.
    """
    print(f"[Step 1] 엑셀 파일 '{filename}' 생성 중...")
    
    # 1. 워크북 생성
    wb = Workbook()
    ws = wb.active # 활성화된 시트 가져오기
    ws.title = "매출보고서" # 시트 이름 변경
    
    # 2. 헤더 추가
    headers = ["날짜", "제품명", "수량", "단가", "총액"]
    ws.append(headers) 
    
    # 3. 헤더 스타일 적용 (진하게, 중앙 정렬, 노란 배경)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
    # 4. 데이터 추가
    data = [
        ["2025-01-01", "키보드", 10, 120000],
        ["2025-01-02", "마우스", 25, 45000],
        ["2025-01-03", "모니터", 5, 350000],
        ["2025-01-04", "노트북", 2, 1500000],
    ]
    
    for row in data:
        # 수량 * 단가 자동 계산하여 추가
        total = row[2] * row[3] 
        ws.append(row + [total])
        
    # 5. 저장
    wb.save(filename)
    print(" -> 엑셀 파일 저장 완료.")

def modify_existing_excel(filename):
    """
    기존 엑셀 파일을 불러와서 수정합니다.
    (예: 총액이 100만원 이상인 경우 빨간색으로 표시)
    """
    print(f"\n[Step 2] 엑셀 파일 '{filename}' 수정 중...")
    
    if not os.path.exists(filename):
        print("파일이 존재하지 않습니다.")
        return

    # 1. 파일 불러오기
    wb = load_workbook(filename)
    ws = wb["매출보고서"]
    
    # 2. 데이터 순회하며 조건 서식 적용
    # min_row=2는 헤더를 제외하고 2번째 줄부터 시작한다는 의미
    for row in ws.iter_rows(min_row=2, max_col=5):
        # row[4]는 '총액' 컬럼 (0부터 시작하므로 인덱스 4)
        total_cell = row[4]
        
        if total_cell.value >= 1000000:
            total_cell.font = Font(color="FF0000", bold=True) # 빨간 글씨
            print(f" -> 고액 매출 발견: {row[1].value} ({total_cell.value}원)")
            
    # 3. 새 이름으로 저장 (원본 보존)
    new_filename = filename.replace(".xlsx", "_updated.xlsx")
    wb.save(new_filename)
    print(f" -> 수정된 파일 저장 완료: {new_filename}")

if __name__ == "__main__":
    file_path = "sales_report_rpa.xlsx"
    
    # 실습 실행
    create_excel_report(file_path)
    modify_existing_excel(file_path)
    
    print("\n[Tip] 생성된 엑셀 파일을 열어서 스타일과 데이터가 잘 들어갔는지 확인해보세요.")
